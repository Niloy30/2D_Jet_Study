# %%
import os
from math import sqrt

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

from obstacle_detection import detect_circles
from surface_perturbation_growth import analyze_steepness_vs_time

g = 9.81

FPS = 7000
experiment_number = "20250514_090243"
results_path = rf"C:\Users\niloy\Google Drive\School Stuff\M.SC Mechanical Engineering\01 - Fluid Dynamics Lab\03 - PDA\01 - 2D Surface Perturbations\Results\{experiment_number}"
excel_sheet  = r"C:\Users\niloy\Google Drive\School Stuff\M.SC Mechanical Engineering\01 - Fluid Dynamics Lab\03 - PDA\01 - 2D Surface Perturbations\Results\results.xlsx"
#experiment_path = rf"C:\Users\niloy\Desktop\Experiments\05022025\{experiment_number}"

conversion_factor = np.load(rf"{results_path}\conversion_factor.npy")
free_surface_all = np.load(rf"{results_path}\free_surface_data.npy")


# if not os.path.exists(rf"{results_path}\obstacle_data.txt"):
#     detect_circles(
#         experiment_path,
#         save_obstacle_data=True,
#         save_path=results_path,
#         show=True
#     )
circle = np.loadtxt(rf"{results_path}\obstacle_data.txt") 
x_circle, y_circle, r_circle = circle * conversion_factor
y_circle = 1024* conversion_factor - y_circle



print(f"Obstacle Position is {y_circle}")

X = free_surface_all[0, :, :]* conversion_factor
Y = free_surface_all[1, :, :]* conversion_factor
T = np.arange(0, X.shape[1], 1)/FPS

Y_average = np.mean(Y, axis=0) - y_circle 
Y_init = np.min(Y_average)
y_circle = 0


# image = np.flipud(cv2.imread(rf'C:\Users\niloy\Desktop\Experiments\04302025\20250430_125933\20250430_125933000002.bmp', cv2.IMREAD_COLOR))

#  plt.imshow(image, origin="lower")
# plt.hlines(np.min(Y_average), 0,1024)
# plt.hlines(y_circle,0,1024,linestyles="-", colors="b", label="Obstacle Center")

# plt.hlines((y_circle+r_circle),0,1024,linestyles="--", colors="b", 
#            label="Obstacle top")
# plt.legend()
# plt.show()
plt.plot(T,Y_average,"blue")
# %%

H0 = np.min(Y_average)
T0 = np.min(T[(Y_average) > 2*H0])
#T0 = 0.018
try:
    TF = np.min(T[(Y_average) > Y_init+65])
except:
    TF = np.nan
plt.plot(T,Y_average,"blue")
plt.hlines(y_circle,0,max(T),linestyles="-", colors="k", label="Obstacle Center")
plt.hlines(y_circle+r_circle,0,max(T),linestyles="--", colors="k", label="Obstacle top")
plt.hlines(Y_init+65, 0,max(T),'k')
plt.legend()
plt.hlines(Y_init+H0,0,max(T),linestyles="-", colors="k", label="Obstacle Center")
plt.grid(True, which='both', linestyle='--', linewidth=0.5, alpha=0.7)
plt.title(f"R = {r_circle:.2g}")
# Optionally, turn on minor ticks for more lines
plt.minorticks_on()
plt.show()
# %%
U0 = H0/T0
A = (2*U0)/ T0
h = H0/r_circle - 1
t = T / T0
Lambda = 2 + g*H0/U0**2

Fr = 1/sqrt(Lambda)

print(Fr)



# %%
analyze_steepness_vs_time(results_path, FPS,n_obstacle=1)
steepness_data = np.load(rf"{results_path}\steepness_data.npy")
s_smooth = steepness_data[0]
upper = steepness_data[1]
lower = steepness_data[2]
# %%
# plt.figure(figsize=(10, 5))
# plt.plot(T/T0, (Y_average-y_circle)/H0, marker=".", linestyle="--", color="b", label="Avg Y")
# plt.xlabel("Time")
# plt.ylabel("Average Y Position")
# plt.title(f"Shot {experiment_number}")
# plt.grid(True)
# plt.hlines(2, 0,1,linestyles="--",colors="k")
# plt.vlines(1,0,2,linestyles="--",colors="k")
# plt.title(f"$h = {h:.2g}$, Fr$ = {Fr:.2g}$")

# plt.figure(figsize=(10, 4))
# plt.plot(T/T0, s_smooth, label='Steepness $s$', color='blue')
# plt.fill_between(T/T0, lower, upper, color='blue', alpha=0.4, label='RMSE')
# plt.hlines(0.2, 0, np.max(T/T0), linestyles="--", colors="k", label=r'Jet Threshold $s_{\mathrm{br}} = 0.2$')
# plt.xlabel('Time [s]')
# plt.ylabel('Steepness $s$')
# plt.grid(True)
# plt.legend()
# plt.tight_layout()

fig, axs = plt.subplots(2, 1, figsize=(10, 8), sharex=True)

# First subplot (Average Y Position)
axs[0].plot(T/T0, Y_average / H0, marker=".", linestyle="--", color="b", label="Free Surface")
axs[0].set_ylabel("Initial Fill Depth H0")
axs[0].hlines(2, 0, 1, linestyles="--", colors="k")
axs[0].vlines(1, 0, 2, linestyles="--", colors="k")
axs[0].hlines((Y_init + 65) / H0, 0, TF/T0, linestyles="--", colors="r", label = "Piston Limit")
axs[0].vlines(TF/T0, 0, (Y_init + 65) / H0, linestyles="--", colors="r")
axs[0].grid(True, which='both', linestyle='--', linewidth=0.5, alpha=0.7)
axs[0].minorticks_on()
axs[0].set_title(f"$h = {h:.2g}$, Fr$ = {Fr:.2g}$")
axs[0].legend()
# Second subplot (Steepness)
axs[1].plot(T/T0, s_smooth, label='Steepness $s$', color='blue')
axs[1].fill_between(T/T0, lower, upper, color='blue', alpha=0.4, label='RMSE')
axs[1].hlines(0.2, 0, np.max(T/T0), linestyles="--", colors="k", label=r'Jet Threshold $s_{\mathrm{br}} = 0.2$')
axs[1].set_xlabel(r'Time $t^* = \frac{T}{T_0}$')
axs[1].set_ylabel('Steepness $s$')
axs[1].grid(True)
axs[1].legend()

plt.tight_layout()
plt.savefig(rf"{results_path}\result.pdf")
plt.show()

# %%
def append_to_results(file_path, experiment_number, h, Fr):
    # Load the existing workbook
    workbook = load_workbook(file_path)
    
    # Select the active worksheet (you can specify by name with workbook["SheetName"] if needed)
    sheet = workbook.active

    # Find the next available row
    next_row = sheet.max_row + 1

    # Write the data to columns A (1), B (2), C (3)
    sheet.cell(row=next_row, column=1, value=experiment_number)
    sheet.cell(row=next_row, column=2, value=h)
    sheet.cell(row=next_row, column=3, value=Fr)

    # Save the workbook
    workbook.save(file_path)

append_to_results(excel_sheet, experiment_number, h, Fr)
# %%
