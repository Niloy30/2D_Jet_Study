# %%
from math import sqrt

import matplotlib.cm as cm
import matplotlib.colors as mcolors
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from scipy.signal import savgol_filter

from surface_perturbation_growth import analyze_steepness_vs_time

# %%
g = 9.81

FPS = 7000
experiment_number = "20250516_115833"
results_path = rf"C:\Users\niloy\Google Drive\School Stuff\M.SC Mechanical Engineering\01 - Fluid Dynamics Lab\03 - PDA\01 - 2D Surface Perturbations\Results\{experiment_number}"
excel_sheet = r"C:\Users\niloy\Google Drive\School Stuff\M.SC Mechanical Engineering\01 - Fluid Dynamics Lab\03 - PDA\01 - 2D Surface Perturbations\Results\results.xlsx"

conversion_factor = np.load(rf"{results_path}\conversion_factor.npy")
free_surface_all = np.load(rf"{results_path}\free_surface_data.npy")

circle = np.loadtxt(rf"{results_path}\obstacle_data.txt")
x_circle, y_circle, r_circle = circle * conversion_factor
y_circle = 1024 * conversion_factor - y_circle


print(f"Obstacle Position is {y_circle}")

X = free_surface_all[0, :, :] * conversion_factor
Y = free_surface_all[1, :, :] * conversion_factor
T = np.arange(0, X.shape[1], 1) / FPS

Y_average = np.mean(Y, axis=0) - y_circle
Y_init = np.min(Y_average)
y_circle_raw = y_circle
y_circle = 0


# plt.plot(T, Y_average, "blue")
# %%

H0 = np.min(Y_average)
T0 = np.min(T[(Y_average) > 2 * H0])
# T0 = 0.01
try:
    TF = np.min(T[(Y_average) > Y_init + 65])
except:
    TF = np.nan
# plt.plot(T, Y_average, "blue")
# plt.hlines(y_circle, 0, max(T), linestyles="-", colors="k", label="Obstacle Center")
# plt.hlines(
#     y_circle + r_circle, 0, max(T), linestyles="--", colors="k", label="Obstacle top"
# )
# plt.hlines(Y_init + 65, 0, max(T), "k")
# plt.legend()
# plt.hlines(Y_init + H0, 0, max(T), linestyles="-", colors="k", label="Obstacle Center")
# plt.grid(True, which="both", linestyle="--", linewidth=0.5, alpha=0.7)
# plt.title(f"R = {r_circle:.2g}")
# # Optionally, turn on minor ticks for more lines
# plt.minorticks_on()
# plt.show()
# %%
U0 = H0 / T0
A = (2 * U0) / T0
h = H0 / r_circle - 1
t = T / T0
Lambda = 2 + g * H0 / U0**2

Fr = 1 / sqrt(Lambda)

print(Fr)


# %%
analyze_steepness_vs_time(results_path, FPS, n_obstacle=1)
steepness_data = np.load(rf"{results_path}\steepness_data.npy")
s_smooth = steepness_data[0]
upper = steepness_data[1]
lower = steepness_data[2]

# Load critical points
critical_points = np.load(rf"{results_path}\steepness_critical_points.npy")

# plt.plot(T, Y_average, "blue")
# plt.hlines(y_circle, 0, max(T), linestyles="-", colors="k", label="Obstacle Center")
# plt.hlines(
#     y_circle + r_circle, 0, max(T), linestyles="--", colors="k", label="Obstacle top"
# )
# plt.hlines(Y_init + 65, 0, max(T), "k")
# plt.plot(T, (critical_points[0, :, 1] + critical_points[2, :, 1]) / 2 - y_circle_raw)
# plt.plot(T, critical_points[1, :, 1] - y_circle_raw)
# plt.legend()
# plt.hlines(Y_init + H0, 0, max(T), linestyles="-", colors="k", label="Obstacle Center")
# plt.grid(True, which="both", linestyle="--", linewidth=0.5, alpha=0.7)
# plt.title(f"R = {r_circle:.2g}")
# # Optionally, turn on minor ticks for more lines
# plt.minorticks_on()
# plt.show()

# plt.plot(T, s_smooth, label="Steepness $s$", color="blue")
# plt.show()

# %%
Frame_max = np.size(T) - 50
cmap = cm.get_cmap("Blues")
norm = mcolors.Normalize(vmin=0, vmax=Frame_max - 1)
colors = [cmap(norm(i)) for i in range(Frame_max)]

plt.figure(figsize=(12, 6))
frame = 10
while frame < Frame_max:
    Y_smooth = savgol_filter(Y[:, frame], window_length=10, polyorder=3)
    plt.plot(X[:, frame] - x_circle, Y_smooth - y_circle_raw, color=colors[frame])

    # Plot critical points for this frame
    left_min = critical_points[0, frame]  # [x, y]
    center = critical_points[1, frame]  # [x, y]
    right_min = critical_points[2, frame]  # [x, y]

    plt.scatter(
        left_min[0] - x_circle,
        left_min[1] - y_circle_raw,
        color="blue",
        marker="o",
        s=30,
        label="Local Min" if frame == 0 else "",
    )
    plt.scatter(
        # center[0] - x_circle,
        0,
        center[1] - y_circle_raw,
        color="red",
        marker="o",
        s=40,
        label="Local Max" if frame == 0 else "",
    )
    plt.scatter(
        right_min[0] - x_circle,
        right_min[1] - y_circle_raw,
        color="blue",
        marker="o",
        s=30,
        # label="Right Min" if frame == 0 else "",
    )

    frame += 10
circle = plt.Circle((0, 0), r_circle, color="black", fill=False)
ax = plt.gca()
ax.add_patch(circle)
ax.set_aspect("equal")
plt.xlim(-50, 50)
# plt.title("Free Surface Evolution")
plt.xlabel("$x$ [mm]")
plt.ylabel("$y$ [mm]")
plt.tight_layout()
# plt.legend(loc="upper right")
plt.tight_layout()
plt.savefig(rf"{results_path}\free_surface_evolution.pdf")
plt.show()
# %%

fig, axs = plt.subplots(2, 1, figsize=(10, 8), sharex=True)

# First subplot (Average Y Position)
axs[0].plot(T, Y_average, color="b", label="Free Surface")
axs[0].set_ylabel("Free surface Height [mm]")
axs[0].plot(
    T,
    (critical_points[0, :, 1] + critical_points[2, :, 1]) / 2 - y_circle_raw,
    color="g",
    label="Local Minima",
)
axs[0].plot(T, critical_points[1, :, 1] - y_circle_raw, color="r", label="Local Maxima")

axs[0].hlines((Y_init + 65), 0, TF, linestyles="--", colors="r", label="Piston Limit")
axs[0].vlines(TF, 0, (Y_init + 65), linestyles="--", colors="r")
# axs[0].grid(True, which="both", linestyle="--", linewidth=0.5, alpha=0.7)
axs[0].minorticks_on()
# axs[0].set_title(f"$h = {h:.2g}$, Fr$ = {Fr:.2g}$")
axs[0].legend()

# Second subplot (Steepness)
axs[1].plot(T, s_smooth, label="Steepness $s$", color="blue")
axs[1].fill_between(T, lower, upper, color="blue", alpha=0.4, label="RMSE")
axs[1].hlines(
    0.2,
    0,
    np.max(T),
    linestyles="--",
    colors="k",
    label=r"Jet Threshold $s_{\mathrm{br}} = 0.2$",
)
axs[1].set_xlabel(r"$t$ [s]")
axs[1].set_ylabel("Steepness $s$")
# axs[1].grid(True)
axs[1].legend()

plt.tight_layout()
# plt.savefig(rf"{results_path}\result_dimensional.pdf")
plt.show()
# %%
fig, axs = plt.subplots(2, 1, figsize=(10, 8), sharex=True)

# First subplot (Average Y Position)
axs[0].plot(
    T / T0, Y_average / H0, marker=".", linestyle="--", color="b", label="Free Surface"
)
axs[0].set_ylabel("Initial Fill Depth H0")
axs[0].hlines(2, 0, 1, linestyles="--", colors="k")
axs[0].vlines(1, 0, 2, linestyles="--", colors="k")
axs[0].hlines(
    (Y_init + 65) / H0, 0, TF / T0, linestyles="--", colors="r", label="Piston Limit"
)
axs[0].vlines(TF / T0, 0, (Y_init + 65) / H0, linestyles="--", colors="r")
# axs[0].grid(True, which="both", linestyle="--", linewidth=0.5, alpha=0.7)
axs[0].minorticks_on()
axs[0].set_title(f"$h = {h:.2g}$, Fr$ = {Fr:.2g}$")
axs[0].legend()
# Second subplot (Steepness)
axs[1].plot(T / T0, s_smooth, label="Steepness $s$", color="blue")
axs[1].fill_between(T / T0, lower, upper, color="blue", alpha=0.4, label="RMSE")
axs[1].hlines(
    0.2,
    0,
    np.max(T / T0),
    linestyles="--",
    colors="k",
    label=r"Jet Threshold $s_{\mathrm{br}} = 0.2$",
)
axs[1].set_xlabel(r"Time $t^* = \frac{T}{T_0}$")
axs[1].set_ylabel("Steepness $s$")
# axs[1].grid(True)
axs[1].legend()

plt.tight_layout()
plt.savefig(rf"{results_path}\result.pdf")
plt.show()


# %%
def append_to_results(file_path, experiment_number, h, Fr):
    # Load the existing workbook
    workbook = load_workbook(file_path)

    # Select the active worksheet (you can specify by name with workbook["SheetName"] if needed)
    sheet = workbook.active

    # Find the next available row
    next_row = sheet.max_row + 1

    # Write the data to columns A (1), B (2), C (3)
    sheet.cell(row=next_row, column=1, value=experiment_number)
    sheet.cell(row=next_row, column=2, value=h)
    sheet.cell(row=next_row, column=3, value=Fr)

    # Save the workbook
    workbook.save(file_path)
